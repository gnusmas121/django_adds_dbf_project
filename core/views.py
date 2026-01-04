# core/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import connection
from django.contrib import messages
import dbfread
import tempfile
import os
import re # Для проверки имени таблицы
import pandas as pd # Используем pandas для удобного чтения Excel
from django.http import JsonResponse, HttpResponse
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from psycopg2 import sql
from .models import DBFUpload, ExcelUpload, TableTemplate, TableTemplateFieldConfig # Импортируем новую модель

# Вспомогательная функция для проверки, является ли пользователь суперпользователем
def is_superuser(user):
    return user.is_superuser

# Вспомогательная функция для проверки, может ли пользователь использовать поиск
def can_search(user):
    # Проверяем, является ли пользователь суперпользователем ИЛИ состоит в группе 'can_search' ИЛИ is_staff
    # is_staff устанавливается через AUTH_LDAP_USER_FLAGS_BY_GROUP, если пользователь в нужной группе ADDS
    return user.is_superuser or user.groups.filter(name='can_search').exists() or user.is_staff
@login_required # Пользователь должен быть аутентифицирован
@user_passes_test(can_search) # Пользователь должен пройти проверку can_search
def search(request):
    results = []
    available_tables = []

    # --- Получаем список таблиц ---
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT tablename
            FROM pg_tables
            WHERE schemaname = 'public'
              AND tablename NOT LIKE 'pg_%'
              AND tablename NOT LIKE 'sql_%'
              AND tablename NOT LIKE 'django_%'
              AND tablename NOT LIKE 'auth_%'
              AND tablename NOT LIKE 'contenttype_%'
              AND tablename NOT LIKE 'core_%'; -- <-- Исключаем таблицы core
        """)
        all_tables = [row[0] for row in cursor.fetchall()]
        available_tables = all_tables

    # --- Выбираем таблицу для поиска ---
    # ПЕРЕМЕННАЯ ОБЯЗАТЕЛЬНО ОБЪЯВЛЯЕТСЯ ЗДЕСЬ
    table_to_search = request.GET.get('table', '')

    # Проверяем, что выбранная таблица существует в списке
    if table_to_search and table_to_search in available_tables:
        print(f"DEBUG: search view - Processing table: {table_to_search}") # <-- Отладка
        # --- Получаем все столбцы таблицы для формирования условий ---
        with connection.cursor() as cursor:
            # ИСПОЛЬЗУЕМ обычную строку с параметром
            query = "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position;"
            print(f"DEBUG: search view - Executing query for table: {table_to_search}") # <-- Отладка
            cursor.execute(query, [table_to_search])
            all_columns = [row[0] for row in cursor.fetchall()]
            print(f"DEBUG: search view - Fetched columns: {all_columns}") # <-- Отладка

        # --- Собираем значения из формы (GET параметров) для *всех возможных* столбцов ---
        # и только для тех, которые были отправлены
        search_values = {}
        for field_name in all_columns:
            value = request.GET.get(field_name, '')
            if value: # Только если значение введено
                search_values[field_name] = value

        print(f"DEBUG search_values: {search_values}") # <-- Отладка

        # --- Собираем поля для вывода ---
        # Получаем список полей из параметра result_fields
        result_fields = request.GET.getlist('result_fields') # Используем getlist для множественных значений
        # Если не выбраны, выводим все
        if not result_fields:
            result_fields = all_columns

        # Проверяем, что выбранные поля вывода существуют в таблице
        result_fields = [f for f in result_fields if f in all_columns]

        print(f"DEBUG result_fields: {result_fields}") # <-- Отладка

        # Формируем условия WHERE и параметры для SQL, только для заполненных полей
        where_parts = []
        params = []
        for field_name, search_value in search_values.items():
            if search_value: # Только если значение введено
                # ПРЕОБРАЗУЕМ поисковое значение ИЗ UTF-8 В CP866 (если база в cp866)
                try:
                    search_value_cp866 = search_value.encode('cp866').decode('cp866')
                    # Используем ILIKE для поиска части строки (регистронезависимо)
                    # Экранируем имя столбца с помощью двойных кавычек
                    # %s - плейсхолдер для psycopg2
                    where_parts.append(f'"{field_name}" ILIKE %s')
                    # Добавляем значение (уже в cp866) с подстановочными знаками % для поиска части строки
                    params.append(f'%{search_value_cp866}%')
                except UnicodeEncodeError:
                    # Обработка ошибки, если строку нельзя закодировать в cp866
                    print(f"Warning: Could not encode search value '{search_value}' to cp866 for field '{field_name}'. Skipping this field.")
                    continue # Пропускаем это поле в поиске

        print(f"DEBUG where_parts: {where_parts}") # <-- Отладка
        print(f"DEBUG params: {params}")           # <-- Отладка

        # Выполняем поиск, если есть условия (хотя бы одно поле заполнено и закодировалось)
        if where_parts:
            # Составляем SQL-запрос
            # SELECT - выбираем *только* выбранные поля результата (экранируем имена)
            # WHERE - условия, объединённые через AND, только для заполненных и закодированных полей
            select_cols = ', '.join([f'"{col}"' for col in result_fields])
            where_clause = ' AND '.join(where_parts)
            # Экранируем имя таблицы
            sql_query = f'SELECT {select_cols} FROM "{table_to_search}" WHERE {where_clause};'

            print(f"DEBUG SQL Query: {sql_query}") # <-- Отладка: выводим SQL

            with connection.cursor() as cursor:
                # Устанавливаем client_encoding для текущей сессии, если данные в базе в cp866
                cursor.execute("SET client_encoding = 'WIN866';") # Или 'cp866'
                cursor.execute(sql_query, params)
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in rows]
        else:
            print("DEBUG: No conditions for WHERE clause, skipping query execution.")
            pass # Если не заполнены поля или закодировать не удалось, возвращаем пустой результат
    # else: # Необязательно, но логично
    #     table_to_search = None # Уже равно '', но можно явно указать

    # --- Передаём данные в шаблон ---
    # Передаём значения формы для отображения (теперь динамически)
    # search_form_values = {field: request.GET.get(field, '') for field in ['P2', 'ST2', 'DNR']} # <-- Больше не нужно

    return render(request, 'core/search.html', {
        'results': results,
        'available_tables': available_tables,
        'selected_table': table_to_search, # <-- Теперь переменная всегда определена
        # 'search_values': search_form_values, # <-- Больше не нужно
    })

# ... (остальные функции) ...

@login_required
@user_passes_test(is_superuser) # Только суперпользователи
def upload_dbf(request):
    if request.method == 'POST' and request.FILES.get('dbf_file'):
        uploaded_file = request.FILES['dbf_file']
        filename = uploaded_file.name

        # Проверка расширения
        if not filename.lower().endswith('.dbf'):
            # Обработка ошибки: неверный формат файла
            return render(request, 'core/upload_dbf.html', {'error': 'Файл должен быть в формате .dbf'})

        # Получаем имя таблицы из имени файла (без расширения)
        table_name = os.path.splitext(filename)[0]
        print(f"DEBUG: Original filename: {filename}, Derived table_name: {table_name}") # <-- Отладка

        # Проверка имени таблицы на безопасность (только буквы, цифры, подчеркивания)
        if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table_name):
             return render(request, 'core/upload_dbf.html', {'error': 'Имя файла содержит недопустимые символы для имени таблицы.'})

        try:
            # Создаем временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix='.dbf') as temp_file:
                for chunk in uploaded_file.chunks():
                    temp_file.write(chunk)
                temp_file_path = temp_file.name

            # Читаем DBF-файл с кодировкой cp866 (DOS)
            table = dbfread.DBF(temp_file_path, encoding='cp866')
            records = list(table)

            if not records:
                os.unlink(temp_file_path)
                return render(request, 'core/upload_dbf.html', {'error': 'Файл DBF пуст.'})

            # Получаем имена полей из DBF
            field_names = list(records[0].keys())

            # --- Определение максимальных длин строковых полей по ВСЕМ записям и определение типов ---
            max_lengths = {}
            type_hints = {} # Словарь для хранения признака, что поле содержит числа (пока грубо)

            for record in records:
                for field_name, value in record.items():
                    # Определяем тип значения в Python
                    val_type = type(value)
                    str_val = str(value) if value is not None else ""

                    # Обновляем максимальную длину для строк, независимо от "типа" в DBF
                    # Если значение - строка или None, проверяем его длину
                    if isinstance(value, str) or value is None:
                         current_len = len(str_val)
                         if field_name not in max_lengths or current_len > max_lengths[field_name]:
                             max_lengths[field_name] = current_len
                    # Если значение - число, запоминаем это (грубо)
                    elif isinstance(value, (int, float)):
                         # Пока просто отметим, что в этом поле был числовой тип
                         # Это не идеально, но помогает в грубой классификации
                         if field_name not in type_hints:
                             type_hints[field_name] = set()
                         type_hints[field_name].add(val_type.__name__)

            # --- Определение типов полей для SQL ---
            field_types = {}
            for field_name in field_names:
                 # Берём значение из первой записи для грубой типизации
                 first_val = records[0][field_name]
                 first_val_type = type(first_val)

                 # Проверяем, было ли в этом поле числовое значение (грубо)
                 was_numeric = type_hints.get(field_name, set()).intersection({'int', 'float'})

                 # Если первое значение - число (int/float) и в других записях тоже были числа
                 if isinstance(first_val, (int, float)) and was_numeric:
                     if isinstance(first_val, int):
                         field_types[field_name] = 'INTEGER'
                     else: # float
                         field_types[field_name] = 'NUMERIC'
                 else:
                     # Для всех остальных случаев (включая строки, None, и числа, которые не числа в других записях)
                     # Используем строковый тип с максимальной длиной
                     max_len_for_field = max_lengths.get(field_name, 0)
                     final_length = max(255, max_len_for_field) # Минимум 255
                     field_types[field_name] = f'VARCHAR({final_length})'

            # print(f"DEBUG: field_types = {field_types}") # Для отладки
            # print(f"DEBUG: max_lengths = {max_lengths}") # Для отладки
            # print(f"DEBUG: type_hints = {type_hints}") # Для отладки
            # Удаляем временный файл
            os.unlink(temp_file_path)

            # Создаем таблицу в PostgreSQL
            with connection.cursor() as cursor:
                # --- УДАЛЯЕМ ТАБЛИЦУ, ЕСЛИ ОНА СУЩЕСТВУЕТ ---
                drop_sql = f"DROP TABLE IF EXISTS \"{table_name}\";"
                cursor.execute(drop_sql)
                # ---

                # Составляем SQL для создания таблицы
                create_sql_parts = []
                for field_name, field_type in field_types.items():
                    # Экранируем имя поля, на случай если оно совпадает с ключевым словом
                    escaped_field_name = f'"{field_name}"'
                    create_sql_parts.append(f"{escaped_field_name} {field_type}")

                create_sql = f"CREATE TABLE \"{table_name}\" ({', '.join(create_sql_parts)});"
                cursor.execute(create_sql)

                # Составляем SQL для вставки данных
                if records:
                    # Подготавливаем столбцы
                    columns = ', '.join([f'"{name}"' for name in field_names])
                    # Подготавливаем плейсхолдеры для значений
                    placeholders = ', '.join(['%s'] * len(field_names))
                    insert_sql = f"INSERT INTO \"{table_name}\" ({columns}) VALUES ({placeholders});"

                    # Подготавливаем данные для вставки
                    data_to_insert = []
                    for record in records:
                        row_data = [record[field_name] for field_name in field_names]
                        data_to_insert.append(tuple(row_data))

                    # Выполняем вставку
                    cursor.executemany(insert_sql, data_to_insert)

            # Сохраняем запись о загрузке (если используем модель)
            # DBFUpload.objects.create(
            #     filename=filename,
            #     table_name=table_name,
            #     uploaded_by=request.user
            # )

            # Успешно
            return redirect('core:search') # Перенаправляем на страницу поиска или другую

        except Exception as e:
            # Удаляем временный файл в случае ошибки
            if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            # Обработка ошибки
            return render(request, 'core/upload_dbf.html', {'error': f'Ошибка обработки файла: {str(e)}'})

    return render(request, 'core/upload_dbf.html')

# --- НОВАЯ ФУНКЦИЯ ДЛЯ ЗАГРУЗКИ EXCEL ---
@login_required
@user_passes_test(is_superuser) # Только суперпользователи
def upload_excel(request):
    print("DEBUG: upload_excel view entered") # <-- Отладка
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        filename = excel_file.name
        print(f"DEBUG: Received file: {filename}") # <-- Отладка

        if not filename.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Пожалуйста, загрузите файл Excel (.xlsx или .xls).')
            return redirect('core:upload_excel')

        # Получаем имя таблицы из имени файла (без расширения)
        table_name = os.path.splitext(filename)[0]
        print(f"DEBUG: Derived table_name: {table_name}") # <-- Отладка

        # Проверка имени таблицы на безопасность (только буквы, цифры, подчеркивания)
        if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table_name):
             messages.error(request, 'Имя файла содержит недопустимые символы для имени таблицы.')
             print(f"DEBUG: Invalid table name: {table_name}") # <-- Отладка
             return redirect('core:upload_excel')

        try:
            # Читаем Excel файл с помощью pandas
            # force_strings=True заставляет pandas читать ВСЁ как строки
            print(f"DEBUG: Attempting to read file with pandas (force_strings=True): {filename}") # <-- Отладка
            df = pd.read_excel(excel_file, engine='openpyxl' if filename.endswith('.xlsx') else 'xlrd', dtype=str)
            print(f"DEBUG: Successfully read file. DataFrame shape: {df.shape}") # <-- Отладка
            print(f"DEBUG: DataFrame head (first 5 rows):\n{df.head()}") # <-- Отладка: первые 5 строк
            print(f"DEBUG: DataFrame columns: {df.columns.tolist()}") # <-- Отладка: имена колонок
            print(f"DEBUG: DataFrame dtypes:\n{df.dtypes}") # <-- Отладка: типы данных после force_strings

            if df.empty:
                messages.error(request, f'Файл Excel {filename} пуст (DataFrame пуст).')
                print(f"DEBUG: DataFrame is empty") # <-- Отладка
                return redirect('core:upload_excel')

            # --- Определение типов полей для SQL: ВСЕ как VARCHAR ---
            # Так как dtype=str в read_excel, все колонки будут object (str в pandas)
            field_types = {}
            for col_name in df.columns:
                # Определяем максимальную длину строки в колонке
                # Преобразуем в строку (на случай, если в данных были NaN, которые стали 'nan')
                max_len_series = df[col_name].astype(str).str.len()
                # Находим максимальное значение длины
                max_len = max_len_series.max()
                # Учитываем, что NaN может стать строкой 'nan' длиной 3, берем длину NaN как 3
                # Устанавливаем минимальную длину 255, как в DBF
                final_length = max(255, max_len if pd.notna(max_len) else 3)
                field_types[col_name] = f'VARCHAR({final_length})'

            print(f"DEBUG: Final field_types (all VARCHAR): {field_types}") # <-- Отладка

            # --- Создание таблицы в PostgreSQL ---
            with connection.cursor() as cursor:
                print(f"DEBUG: Attempting to drop table: {table_name}") # <-- Отладка
                # --- УДАЛЯЕМ ТАБЛИЦУ, ЕСЛИ ОНА СУЩЕСТВУЕТ ---
                drop_sql = f"DROP TABLE IF EXISTS \"{table_name}\";"
                cursor.execute(drop_sql)
                print(f"DEBUG: Dropped table: {table_name}") # <-- Отладка

                # Составляем SQL для создания таблицы
                create_sql_parts = []
                for field_name, field_type in field_types.items():
                    escaped_field_name = f'"{field_name}"'
                    create_sql_parts.append(f"{escaped_field_name} {field_type}")

                create_sql = f"CREATE TABLE \"{table_name}\" ({', '.join(create_sql_parts)});"
                print(f"DEBUG: Creating table with SQL: {create_sql}") # <-- Отладка
                cursor.execute(create_sql)
                print(f"DEBUG: Created table: {table_name}") # <-- Отладка

                # --- Подготовка и вставка данных ---
                # Все данные уже строки (dtype=str), но нужно обработать NaN, которые стали 'nan', 'None', 'NaT'
                # Заменим их на None (NULL в SQL)
                df_for_insert = df.where(pd.notna(df), None)

                # Преобразуем в список кортежей
                records_to_insert = [tuple(row) for row in df_for_insert.values]
                print(f"DEBUG: Prepared {len(records_to_insert)} records for insertion (values: {df_for_insert.values[:5]})") # <-- Отладка: первые 5 записей

                if records_to_insert:
                    columns = ', '.join([f'"{name}"' for name in df.columns])
                    placeholders = ', '.join(['%s'] * len(df.columns))
                    insert_sql = f'INSERT INTO "{table_name}" ({columns}) VALUES ({placeholders});'

                    print(f"DEBUG: Attempting to insert records into {table_name}") # <-- Отладка
                    # Устанавливаем client_encoding для текущей сессии, если данные в базе в cp866
                    cursor.execute("SET client_encoding = 'WIN866';") # Или 'cp866'
                    cursor.executemany(insert_sql, records_to_insert)
                    print(f"DEBUG: Successfully inserted {len(records_to_insert)} records") # <-- Отладка

            records_count = len(records_to_insert)

            # --- Создание записи о загрузке Excel ---
            ExcelUpload.objects.create(
                filename=filename,
                table_name=table_name,
                uploaded_by=request.user
            )
            print(f"DEBUG: Created ExcelUpload record for {filename} -> {table_name}") # <-- Отладка

            messages.success(request, f'Успешно создана таблица "{table_name}" и загружено {records_count} записей из {filename} как строки.')

        except Exception as e:
            print(f"DEBUG: Exception occurred: {str(e)}") # <-- Отладка
            messages.error(request, f'Ошибка при обработке файла: {str(e)}')

        return redirect('core:upload_excel')

    # Если GET запрос, просто отображаем страницу
    print("DEBUG: GET request, rendering page") # <-- Отладка
    context = {}
    return render(request, 'core/upload_excel.html', context)


@login_required
@user_passes_test(can_search) # Используем существующую проверку, или измените на is_superuser
def download_search_template(request):
    """
    Скачивание Excel-шаблона (с заголовками столбцов) для выбранной таблицы.
    """
    table_name = request.GET.get('table_name')
    if not table_name:
        # Обработка ошибки: имя таблицы не передано
        messages.error(request, 'Не указана таблица для формирования шаблона.')
        return HttpResponse("Не указана таблица", status=400)

    # Проверяем, что имя таблицы "безопасно", чтобы избежать SQL-инъекции
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table_name):
        messages.error(request, 'Недопустимое имя таблицы.')
        return HttpResponse("Недопустимое имя таблицы", status=400)

    # Получаем структуру таблицы (имена столбцов)
    try:
        with connection.cursor() as cursor:
            query = sql.SQL("SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position;")
            cursor.execute(query, [table_name])
            columns = [row[0] for row in cursor.fetchall()]

        if not columns:
            messages.error(request, f'Таблица "{table_name}" не содержит столбцов.')
            return HttpResponse(f'Таблица "{table_name}" не содержит столбцов.', status=404)

    except Exception as e:
        messages.error(request, f'Ошибка при получении структуры таблицы: {str(e)}')
        return HttpResponse(f'Ошибка при получении структуры таблицы: {str(e)}', status=500)

    # Создаём Excel-файл в памяти
    wb = Workbook()

    # --- Лист 1: Шаблон с заголовками ---
    ws_template = wb.active
    ws_template.title = f"Шаблон_{table_name}"

    # Записываем имена столбцов в первую строку (это и есть подписи)
    for col_num, column_title in enumerate(columns, 1):
        cell = ws_template.cell(row=1, column=col_num, value=column_title)
        # Опционально: стиль заголовка
        # cell.font = Font(bold=True)

    # --- Лист 2: Пустой лист для данных ---
    ws_data = wb.create_sheet(title=f"Данные_{table_name}")
    # Оставляем его пустым

    # Устанавливаем ширину столбцов (опционально, для обоих листов)
    for ws in [ws_template, ws_data]:
        for column in ws.columns:
            max_length = 0
            column_letter_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter_name].width = adjusted_width

    # Сохраняем в BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Формируем имя файла
    filename = f"{table_name}_search_template.xlsx"

    # Возвращаем файл как HTTP-ответ
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(can_search)
def get_table_columns(request):
    """
    Возвращает JSON с именами столбцов, их конфигурацией для поиска и вывода для указанной таблицы.
    """
    table_name = request.GET.get('table_name')
    if not table_name:
        return JsonResponse({'error': 'Table name is required'}, status=400)

    # Проверяем, что имя таблицы "безопасно", чтобы избежать SQL-инъекции
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table_name):
        return JsonResponse({'error': 'Invalid table name'}, status=400)

    try:
        with connection.cursor() as cursor:
            # ИСПОЛЬЗУЕМ обычную строку с параметром, как в manage_table_template
            query = "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position;"
            print(f"DEBUG: Executing query for table (get_table_columns): '{table_name}'") # <-- Отладка
            cursor.execute(query, [table_name])
            columns = [row[0] for row in cursor.fetchall()]
            print(f"DEBUG: Fetched columns (get_table_columns): {columns}") # <-- Отладка

        # Проверяем, есть ли шаблон для этой таблицы
        try:
            template_obj = TableTemplate.objects.prefetch_related('field_configs').get(table_name=table_name)
            # Получаем настройки полей из связанной модели
            field_configs = template_obj.field_configs.all()

            # Разделяем настройки по типу
            search_configs = {cfg.field_name: cfg.field_label for cfg in field_configs if cfg.template_type == 'search'}
            result_configs = {cfg.field_name: cfg.field_label for cfg in field_configs if cfg.template_type == 'result'}

            # Получаем порядок из связанной модели, отсортированной по 'order'
            search_order = [cfg.field_name for cfg in field_configs if cfg.template_type == 'search']
            result_order = [cfg.field_name for cfg in field_configs if cfg.template_type == 'result']

            # Если порядок не задан через Inline (например, все поля имеют order=0),
            # используем порядок из базы данных для отсутствующих в шаблоне
            if not search_order:
                search_order = columns
            else:
                # Добавляем столбцы, которые не были в шаблоне, в конец списка
                search_order.extend([col for col in columns if col not in search_order])

            if not result_order:
                result_order = columns
            else:
                # Добавляем столбцы, которые не были в шаблоне, в конец списка
                result_order.extend([col for col in columns if col not in result_order])


            # Подписи для поиска: используем настройку, если есть, иначе имя столбца
            search_labels = {col: search_configs.get(col, col) for col in search_order if col in columns}
            # Подписи для вывода: используем настройку, если есть, иначе имя столбца
            result_labels = {col: result_configs.get(col, col) for col in result_order if col in columns}

        except TableTemplate.DoesNotExist:
            # Если шаблона нет, используем имена столбцов как есть
            search_labels = {col: col for col in columns}
            result_labels = {col: col for col in columns}
            search_order = columns
            result_order = columns

        return JsonResponse({
            'columns': columns,
            'search_labels': search_labels,
            'result_labels': result_labels,
            'search_order': search_order,
            'result_order': result_order
        })
    except Exception as e:
        print(f"DEBUG: Database error in get_table_columns for table '{table_name}': {str(e)}") # <-- Отладка
        return JsonResponse({'error': f'Database error: {str(e)}'}, status=500)


# ... (остальные функции, если есть) ...
@login_required
@user_passes_test(is_superuser) # Только суперпользователь может управлять шаблонами
def manage_table_template(request, table_name=None):
    """
    Страница для выбора таблицы и настройки шаблона (поля поиска/вывода и подписи).
    """
    available_tables = []
    table_columns = []
    template_exists = False
    existing_configs = []
    existing_search_fields = []
    existing_result_fields = []

    # Получаем список таблиц
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT tablename
            FROM pg_tables
            WHERE schemaname = 'public'
              AND tablename NOT LIKE 'pg_%'
              AND tablename NOT LIKE 'sql_%'
              AND tablename NOT LIKE 'django_%'
              AND tablename NOT LIKE 'auth_%'
              AND tablename NOT LIKE 'contenttype_%'
              AND tablename NOT LIKE 'core_%';
        """)
        all_tables = [row[0] for row in cursor.fetchall()]
        available_tables = all_tables

    if request.method == 'POST':
        table_name = request.POST.get('table_name')
        if table_name and table_name in available_tables:
            # Получаем столбцы выбранной таблицы
            with connection.cursor() as cursor:
                query = "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position;"
                print(f"DEBUG: Executing query for table (POST): {table_name}") # <-- Отладка
                cursor.execute(query, [table_name])
                table_columns = [row[0] for row in cursor.fetchall()]
                print(f"DEBUG: Fetched columns (POST): {table_columns}") # <-- Отладка

            # Обработка сохранения
            # Удаляем старые настройки для этой таблицы
            TableTemplateFieldConfig.objects.filter(table_template__table_name=table_name).delete()
            # Удаляем сам шаблон, если он был (чтобы создать заново с новым order)
            TableTemplate.objects.filter(table_name=table_name).delete()

            # Создаём/обновляем шаблон
            template_obj, created = TableTemplate.objects.get_or_create(
                table_name=table_name,
                defaults={'created_by': request.user}
            )
            if not created:
                template_obj.created_by = request.user # Обновляем автора, если нужно
                template_obj.save()

            # --- Обработка полей поиска ---
            # Получаем все ключи, начинающиеся с 'search_select_'
            search_select_keys = [k for k in request.POST.keys() if k.startswith('search_select_')]
            # Получаем все ключи, начинающиеся с 'search_label_'
            search_label_keys = [k for k in request.POST.keys() if k.startswith('search_label_')]

            # Сопоставляем по ID (часть после префикса)
            search_configs = {}
            for key in search_select_keys:
                unique_id = key[len('search_select_'):]
                field_name = request.POST.get(key)
                # Проверяем, есть ли соответствующая подпись для этого ID
                label_key = f'search_label_{unique_id}'
                if label_key in search_label_keys:
                    label = request.POST.get(label_key, field_name) # Если подпись пуста, используем имя поля
                    # Проверяем, что поле существует в таблице
                    if field_name and field_name in table_columns:
                        search_configs[unique_id] = {'field': field_name, 'label': label}

            # Сохраняем поля поиска
            for idx, (unique_id, config_data) in enumerate(search_configs.items()):
                TableTemplateFieldConfig.objects.create(
                    table_template=template_obj,
                    field_name=config_data['field'],
                    field_label=config_data['label'],
                    template_type='search',
                    order=idx # Устанавливаем порядок
                )

            # --- Обработка полей вывода ---
            # Аналогично для result_fields
            result_select_keys = [k for k in request.POST.keys() if k.startswith('result_select_')]
            result_label_keys = [k for k in request.POST.keys() if k.startswith('result_label_')]

            result_configs = {}
            for key in result_select_keys:
                unique_id = key[len('result_select_'):]
                field_name = request.POST.get(key)
                # Проверяем, есть ли соответствующая подпись для этого ID
                label_key = f'result_label_{unique_id}'
                if label_key in result_label_keys:
                    label = request.POST.get(label_key, field_name)
                    # Проверяем, что поле существует в таблице
                    if field_name and field_name in table_columns:
                        result_configs[unique_id] = {'field': field_name, 'label': label}

            # Сохраняем поля вывода
            for idx, (unique_id, config_data) in enumerate(result_configs.items()):
                TableTemplateFieldConfig.objects.create(
                    table_template=template_obj,
                    field_name=config_data['field'],
                    field_label=config_data['label'],
                    template_type='result',
                    order=idx # Устанавливаем порядок
                )

            messages.success(request, f'Шаблон для таблицы "{table_name}" успешно сохранён.')
            # Перенаправляем, чтобы избежать повторной отправки формы при обновлении страницы
            return redirect('core:manage_table_template_with_table', table_name=table_name)

    elif request.GET.get('table_name'): # GET запрос с указанным table_name
        table_name = request.GET.get('table_name')
        if table_name in available_tables:
            # Получаем столбцы выбранной таблицы
            with connection.cursor() as cursor:
                query = "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position;"
                print(f"DEBUG: Executing query for table (GET): {table_name}") # <-- Отладка
                cursor.execute(query, [table_name])
                table_columns = [row[0] for row in cursor.fetchall()]
                print(f"DEBUG: Fetched columns (GET): {table_columns}") # <-- Отладка

            # Проверяем, есть ли шаблон
            try:
                template_obj = TableTemplate.objects.prefetch_related('field_configs').get(table_name=table_name)
                template_exists = True
                existing_configs = template_obj.field_configs.all()

                # Подготовим списки существующих полей для поиска и вывода
                existing_search_fields = [cfg.field_name for cfg in existing_configs if cfg.template_type == 'search']
                existing_result_fields = [cfg.field_name for cfg in existing_configs if cfg.template_type == 'result']

            except TableTemplate.DoesNotExist:
                pass
        else:
            messages.error(request, f'Таблица "{table_name}" не найдена.')
            table_name = None # Сбросим, если таблица не найдена

    # Подготовим данные для шаблона
    context = {
        'available_tables': available_tables,
        'selected_table': table_name,
        'table_columns': table_columns,
        'template_exists': template_exists,
        'existing_configs': existing_configs,
        'existing_search_fields': existing_search_fields, # Передаём в шаблон
        'existing_result_fields': existing_result_fields, # Передаём в шаблон
    }
    return render(request, 'core/manage_table_template.html', context)